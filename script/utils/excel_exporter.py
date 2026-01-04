"""
Excel导出模块
将统计信息导出为格式化的Excel文件
"""
from typing import Any, Dict, Optional

try:
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import (
        Alignment,
        Border,
        Font,
        PatternFill,
        Side,
    )
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

from utils.accuracy_calculator import calculate_accuracy_value
from utils.constants import (
    FIELD_CORRECT,
    FIELD_MCQ_ACCURACY,
    FIELD_MCQ_CORRECT,
    FIELD_MCQ_TOTAL,
    FIELD_NUMERIC_ACCURACY,
    FIELD_NUMERIC_CORRECT,
    FIELD_NUMERIC_TOTAL,
    FIELD_TOTAL,
    FIELD_YESNO_ACCURACY,
    FIELD_YESNO_CORRECT,
    FIELD_YESNO_TOTAL,
)
from utils.formatter import format_stat_string


def export_to_excel(
    stats_by_path: Dict[str, Dict[str, Any]],
    excel_file: Optional[str] = None
) -> Optional[str]:
    """
    将统计信息导出为Excel文件
    
    Args:
        stats_by_path: 按路径类型的统计字典
        excel_file: Excel文件路径，如果为None则自动生成
        
    Returns:
        Excel文件路径，如果导出失败则返回None
    """
    if not EXCEL_AVAILABLE:
        print(
            "警告: 未安装 pandas 或 openpyxl，无法导出Excel文件。"
            "请运行: pip install pandas openpyxl"
        )
        return None
    
    # 计算总计
    totals = {
        FIELD_CORRECT: 0,
        FIELD_TOTAL: 0,
        FIELD_MCQ_CORRECT: 0,
        FIELD_MCQ_TOTAL: 0,
        FIELD_YESNO_CORRECT: 0,
        FIELD_YESNO_TOTAL: 0,
        FIELD_NUMERIC_CORRECT: 0,
        FIELD_NUMERIC_TOTAL: 0
    }
    
    # 准备数据行
    data_rows = []
    for path_type in sorted(stats_by_path.keys()):
        stats = stats_by_path[path_type]
        
        # 累计总计
        for key in totals:
            totals[key] += stats[key]
        
        # 格式化数据
        mcq_str = format_stat_string(
            stats[FIELD_MCQ_CORRECT],
            stats[FIELD_MCQ_TOTAL],
            stats.get(FIELD_MCQ_ACCURACY, 0.0)
        )
        yesno_str = format_stat_string(
            stats[FIELD_YESNO_CORRECT],
            stats[FIELD_YESNO_TOTAL],
            stats.get(FIELD_YESNO_ACCURACY, 0.0)
        )
        numeric_str = format_stat_string(
            stats[FIELD_NUMERIC_CORRECT],
            stats[FIELD_NUMERIC_TOTAL],
            stats.get(FIELD_NUMERIC_ACCURACY, 0.0)
        )
        total_str = format_stat_string(
            stats[FIELD_CORRECT],
            stats[FIELD_TOTAL],
            stats.get('accuracy', 0.0)
        )
        
        data_rows.append({
            '增强路径': path_type,
            '选择题': mcq_str,
            'Yes/No': yesno_str,
            '数值': numeric_str,
            '总计': total_str
        })
    
    # 添加总计行
    if totals[FIELD_TOTAL] > 0:
        overall_accuracy = calculate_accuracy_value(
            totals[FIELD_CORRECT], totals[FIELD_TOTAL]
        )
        mcq_overall_accuracy = calculate_accuracy_value(
            totals[FIELD_MCQ_CORRECT], totals[FIELD_MCQ_TOTAL]
        )
        yesno_overall_accuracy = calculate_accuracy_value(
            totals[FIELD_YESNO_CORRECT], totals[FIELD_YESNO_TOTAL]
        )
        numeric_overall_accuracy = calculate_accuracy_value(
            totals[FIELD_NUMERIC_CORRECT], totals[FIELD_NUMERIC_TOTAL]
        )
        
        mcq_str = format_stat_string(
            totals[FIELD_MCQ_CORRECT],
            totals[FIELD_MCQ_TOTAL],
            mcq_overall_accuracy
        )
        yesno_str = format_stat_string(
            totals[FIELD_YESNO_CORRECT],
            totals[FIELD_YESNO_TOTAL],
            yesno_overall_accuracy
        )
        numeric_str = format_stat_string(
            totals[FIELD_NUMERIC_CORRECT],
            totals[FIELD_NUMERIC_TOTAL],
            numeric_overall_accuracy
        )
        total_str = format_stat_string(
            totals[FIELD_CORRECT],
            totals[FIELD_TOTAL],
            overall_accuracy
        )
        
        data_rows.append({
            '增强路径': '总计',
            '选择题': mcq_str,
            'Yes/No': yesno_str,
            '数值': numeric_str,
            '总计': total_str
        })
    
    # 创建DataFrame
    df = pd.DataFrame(data_rows)
    
    # 生成Excel文件路径
    if excel_file is None:
        excel_file = '各增强路径的正确率统计.xlsx'
    
    # 保存为Excel文件
    df.to_excel(
        excel_file, index=False, sheet_name='正确率统计'
    )
    
    # 使用openpyxl美化表格
    wb = load_workbook(excel_file)
    ws = wb.active
    
    # 设置标题样式
    header_fill = PatternFill(
        start_color="366092", end_color="366092", fill_type="solid"
    )
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # 设置边框
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 应用标题样式
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # 设置数据行样式
    data_font = Font(size=11)
    data_alignment = Alignment(
        horizontal="center", vertical="center"
    )
    
    # 应用数据样式
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = thin_border
    
    # 设置列宽
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20
    
    # 设置行高
    ws.row_dimensions[1].height = 25
    for row_num in range(2, ws.max_row + 1):
        ws.row_dimensions[row_num].height = 20
    
    # 保存文件
    wb.save(excel_file)
    print(f"\nExcel文件已创建: {excel_file}")
    return excel_file
